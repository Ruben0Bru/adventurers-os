import os
from openpyxl import load_workbook

def extraer_catalogo_excel(ruta_archivo="plantilla_requisitos.xlsx"):
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(f"[-] Archivo no encontrado: {ruta_archivo}")

    print(f"[Extractor] Iniciando escaneo heurístico de {ruta_archivo}...")
    catalogo_limpio = []
    
    workbook = load_workbook(filename=ruta_archivo, data_only=True)
    sheet = workbook.active  # Toma la hoja que quedó activa al guardar

    # 1. Búsqueda de Cabeceras (Radar en las primeras 10 filas)
    cabeceras = {}
    fila_cabeceras_idx = None

    # Escaneamos buscando la firma de nuestro contrato
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        temp_headers = {}
        for col_idx, cell in enumerate(row):
            # Guardamos cualquier texto que encontremos en esta fila
            if cell.value and isinstance(cell.value, str):
                temp_headers[cell.value.strip()] = col_idx
        
        # Si la fila tiene al menos estas dos columnas clave, asumimos que ES la cabecera
        if 'Eje_Tematico' in temp_headers and 'Requisito' in temp_headers:
            cabeceras = temp_headers
            fila_cabeceras_idx = row_idx
            break

    # Si terminamos el radar y no encontramos el contrato:
    if not fila_cabeceras_idx:
        workbook.close()
        raise ValueError("[-] Contrato roto. El radar revisó 10 filas y no encontró las columnas 'Eje_Tematico' ni 'Requisito'. ¿Estás en la hoja correcta?")
    
    print(f"[Extractor] Cabeceras detectadas en la fila {fila_cabeceras_idx}. Mapeo: {list(cabeceras.keys())}")

    # 2. Extracción de datos (Desde la fila siguiente a las cabeceras)
    for row in sheet.iter_rows(min_row=fila_cabeceras_idx + 1, values_only=True):
        # Usamos el diccionario de cabeceras para buscar la columna exacta
        eje = str(row[cabeceras.get('Eje_Tematico')]).strip() if cabeceras.get('Eje_Tematico') is not None and row[cabeceras.get('Eje_Tematico')] is not None else ''
        req = str(row[cabeceras.get('Requisito')]).strip() if cabeceras.get('Requisito') is not None and row[cabeceras.get('Requisito')] is not None else ''
        det = str(row[cabeceras.get('Detalle')]).strip() if cabeceras.get('Detalle') is not None and row[cabeceras.get('Detalle')] is not None else ''
        
        # Ignoramos celdas vacías o filas basura
        if req and req.lower() != 'none': 
            catalogo_limpio.append({
                "eje_curricular": eje,
                "titulo": req,
                "descripcion": det
            })
            
    workbook.close()
    print(f"[Extractor] {len(catalogo_limpio)} registros parseados correctamente (UTF-8 garantizado).")
    return catalogo_limpio

if __name__ == "__main__":
    datos = extraer_catalogo_excel()
    print(datos[:2])